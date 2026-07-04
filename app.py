import io
import re
import time
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings("ignore")


# ==================== PAGE CONFIG ====================
st.set_page_config(
    page_title="Ultimate Pro Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ==================== SESSION STATE ====================
defaults = {
    "data": None,
    "clean_data": None,
    "refresh_count": 0,
    "theme_color": "#667eea",
    "clean_log": [],
    "original_dtypes": {},
    "source_sheet": None,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ==================== EXCEL ENGINE ====================

TARGET_SHEETS = {"RM", "PM", "FG", "SFG", "SFG 1"}


def _normalize_label(value):
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _std_col(text):
    text = _normalize_label(text)
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def _extract_day_number(text):
    if not text:
        return None
    t = str(text).strip().lower()
    m = re.search(r"\bday\s*[-_]?\s*(\d{1,2})\b", t)
    if m:
        day = int(m.group(1))
        return day if 1 <= day <= 31 else None
    if re.fullmatch(r"\d{1,2}", t):
        day = int(t)
        return day if 1 <= day <= 31 else None
    return None


def _unique_columns(cols):
    seen = {}
    out = []
    for c in cols:
        base = str(c) if c not in (None, "") else "COLUMN"
        if base not in seen:
            seen[base] = 0
            out.append(base)
        else:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
    return out


def _score_header_row(row_values):
    values = [_normalize_label(v) for v in row_values]
    non_empty = sum(1 for v in values if v)
    if non_empty == 0:
        return -1

    day_hits = sum(1 for v in values if _extract_day_number(v) is not None)
    keyword_hits = sum(
        1
        for v in values
        if re.search(r"\b(in|out|cl|stk|stock|item|code|material|qty|opening|closing)\b", v.lower())
    )
    short_tokens = sum(1 for v in values if 1 <= len(v) <= 20)
    unique_ratio = len(set(v for v in values if v)) / max(non_empty, 1)

    return non_empty + (day_hits * 2) + keyword_hits + short_tokens * 0.2 + unique_ratio


def _detect_header_rows(df_raw):
    max_scan = min(len(df_raw), 20)
    best_idx = 0
    best_score = -1

    for i in range(max_scan):
        score = _score_header_row(df_raw.iloc[i].tolist())
        if score > best_score:
            best_score = score
            best_idx = i

    two_row = False
    if best_idx + 1 < len(df_raw):
        row1 = [_normalize_label(v) for v in df_raw.iloc[best_idx].tolist()]
        row2 = [_normalize_label(v) for v in df_raw.iloc[best_idx + 1].tolist()]

        row1_non_empty = sum(1 for v in row1 if v)
        row2_non_empty = sum(1 for v in row2 if v)
        row1_day = sum(1 for v in row1 if _extract_day_number(v) is not None)
        row2_metric = sum(
            1 for v in row2 if re.search(r"\b(in|out|cl\s*stk|cl|stock|qty)\b", v.lower())
        )

        if row1_non_empty > 0 and row2_non_empty > 0 and (row1_day > 0 or row2_metric >= 2):
            two_row = True

    return best_idx, two_row


def _compose_header(parent, child, idx):
    p = _std_col(parent)
    c = _std_col(child)
    day_num = _extract_day_number(parent)

    if day_num is not None and c:
        return f"DAY_{day_num}_{c}"
    if p and c:
        if p == c:
            return p
        return f"{p}_{c}"
    if c:
        return c
    if p:
        return p
    return f"COL_{idx + 1}"


def _prepare_excel_sheet(df_raw):
    # Fill merged header cells horizontally before header detection.
    top = df_raw.head(5).copy()
    top = top.ffill(axis=1)
    df_raw = pd.concat([top, df_raw.iloc[5:]], ignore_index=True)

    header_idx, two_row_header = _detect_header_rows(df_raw)

    if two_row_header and header_idx + 1 < len(df_raw):
        parent = df_raw.iloc[header_idx].ffill().tolist()
        child = df_raw.iloc[header_idx + 1].tolist()
        columns = [_compose_header(parent[i], child[i], i) for i in range(len(parent))]
        data_start = header_idx + 2
    else:
        base = df_raw.iloc[header_idx].tolist()
        columns = [_compose_header(base[i], None, i) for i in range(len(base))]
        data_start = header_idx + 1

    cleaned = df_raw.iloc[data_start:].copy()
    cleaned.columns = _unique_columns(columns)

    cleaned = cleaned.dropna(how="all")
    cleaned = cleaned.dropna(axis=1, how="all")

    cleaned.columns = _unique_columns([_std_col(c) or f"COL_{i+1}" for i, c in enumerate(cleaned.columns)])

    return cleaned.reset_index(drop=True), header_idx, two_row_header


@st.cache_data(show_spinner=False)
def parse_excel_bytes(file_content, sheet_name):
    raw_df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name, header=None, engine="openpyxl")
    parsed_df, header_idx, two_row = _prepare_excel_sheet(raw_df)
    return parsed_df, header_idx, two_row


def detect_and_fix_column_types(df):
    log = []
    df = df.copy()

    for col in df.columns:
        if isinstance(df[col], pd.DataFrame):
            continue

        original_dtype = str(df[col].dtype)
        sample = df[col].dropna().astype(str).head(100)

        if df[col].dtype == "object":
            bool_map = {
                "yes": True,
                "no": False,
                "true": True,
                "false": False,
                "y": True,
                "n": False,
                "1": True,
                "0": False,
                "active": True,
                "inactive": False,
            }
            lower_vals = sample.str.lower().unique()
            if len(lower_vals) > 0 and all(v in bool_map for v in lower_vals):
                df[col] = df[col].astype(str).str.lower().map(bool_map)
                log.append(f"✅ **{col}**: Converted to Boolean ({original_dtype} → bool)")
                continue

            currency_pattern = sample.str.match(r"^[\$€£₹₦₩¥]?[\s]?[\d,]+\.?\d*[\s]?[kKmMbB]?$")
            if len(sample) and currency_pattern.mean() > 0.7:
                cleaned = (
                    df[col]
                    .astype(str)
                    .str.replace(r"[\$€£₹₦₩¥,\s]", "", regex=True)
                    .str.replace(r"[kK]$", "e3", regex=True)
                    .str.replace(r"[mM]$", "e6", regex=True)
                    .str.replace(r"[bB]$", "e9", regex=True)
                )
                converted = pd.to_numeric(cleaned, errors="coerce")
                if converted.notna().mean() > 0.7:
                    df[col] = converted
                    log.append(f"💰 **{col}**: Currency cleaned & converted to numeric")
                    continue

            pct_pattern = sample.str.match(r"^[\d,]+\.?\d*\s*%$")
            if len(sample) and pct_pattern.mean() > 0.7:
                df[col] = pd.to_numeric(df[col].astype(str).str.replace("%", "").str.strip(), errors="coerce")
                log.append(f"📊 **{col}**: Percentage stripped & converted to numeric")
                continue

            date_formats = [
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%m/%d/%Y",
                "%d-%m-%Y",
                "%m-%d-%Y",
                "%Y/%m/%d",
                "%d %b %Y",
                "%d %B %Y",
                "%b %d, %Y",
                "%B %d, %Y",
                "%Y%m%d",
                "%d-%b-%Y",
                "%d/%b/%Y",
            ]
            date_converted = False
            for fmt in date_formats:
                converted = pd.to_datetime(df[col], format=fmt, errors="coerce")
                if converted.notna().mean() > 0.7:
                    df[col] = converted
                    log.append(f"📅 **{col}**: Converted to Date (format: {fmt})")
                    date_converted = True
                    break

            if date_converted:
                continue

            cleaned_num = df[col].astype(str).str.replace(",", "").str.strip()
            converted = pd.to_numeric(cleaned_num, errors="coerce")
            if converted.notna().mean() > 0.75:
                df[col] = converted
                log.append(f"🔢 **{col}**: Comma-separated numbers cleaned → numeric")
                continue

        if df[col].dtype in ["int64", "float64"]:
            try:
                df[col] = pd.to_numeric(
                    df[col],
                    downcast="integer" if df[col].dtype == "int64" else "float",
                )
            except Exception:
                pass

    return df, log


def remove_outliers_iqr(df, cols, multiplier=3.0):
    log = []
    for col in cols:
        if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
            q1 = df[col].quantile(0.25)
            q3 = df[col].quantile(0.75)
            iqr = q3 - q1
            lower = q1 - multiplier * iqr
            upper = q3 + multiplier * iqr
            before = len(df)
            df = df[(df[col] >= lower) & (df[col] <= upper)]
            removed = before - len(df)
            if removed > 0:
                log.append(f"🎯 **{col}**: Removed {removed} outliers (range: {lower:.2f} – {upper:.2f})")
    return df, log


def standardize_text_columns(df, cols):
    log = []
    na_variants = {
        "n/a",
        "na",
        "none",
        "null",
        "nil",
        "-",
        "--",
        "---",
        "not available",
        "not applicable",
        "nan",
        "missing",
        "#n/a",
        "",
    }

    for col in cols:
        if col not in df.columns or isinstance(df[col], pd.DataFrame):
            continue

        original = df[col].copy()
        df[col] = df[col].astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
        df[col] = df[col].apply(lambda x: np.nan if str(x).strip().lower() in na_variants else x)

        changed = (df[col] != original.astype(str)).sum()
        if changed > 0:
            log.append(f"✏️ **{col}**: Standardized {changed} text values")

    return df, log


def fix_negative_values(df, cols):
    log = []
    for col in cols:
        if col in df.columns and pd.api.types.is_numeric_dtype(df[col]):
            neg_count = (df[col] < 0).sum()
            if neg_count > 0:
                log.append(f"⚠️ **{col}**: {neg_count} negative values found (min: {df[col].min():,.2f})")
    return df, log


def deep_clean_data(df, options):
    all_logs = []
    original_rows = len(df)
    original_cols = len(df.columns)

    if df.columns.duplicated().any():
        cols = pd.Series(df.columns)
        for dup in cols[cols.duplicated()].unique():
            count = 0
            for i, c in enumerate(cols):
                if c == dup:
                    if count > 0:
                        cols.iloc[i] = f"{c}_{count}"
                    count += 1
        df.columns = cols
        all_logs.append("🔧 **Fixed**: Duplicate column names resolved")

    if options.get("empty_rows", True):
        before_r = len(df)
        df = df.dropna(how="all")
        removed_r = before_r - len(df)
        if removed_r > 0:
            all_logs.append(f"🗑️ Removed {removed_r} completely empty rows")

    if options.get("empty_cols", True):
        before_c = len(df.columns)
        df = df.dropna(axis=1, how="all")
        removed_c = before_c - len(df.columns)
        if removed_c > 0:
            all_logs.append(f"🗑️ Removed {removed_c} completely empty columns")

    if options.get("standardize", True):
        df.columns = (
            df.columns.astype(str)
            .str.strip()
            .str.upper()
            .str.replace(r"[^\w]", "_", regex=True)
            .str.replace(r"_+", "_", regex=True)
            .str.strip("_")
        )
        all_logs.append("📝 Column names standardized (UPPER_SNAKE_CASE)")

    if options.get("smart_types", True):
        df, type_logs = detect_and_fix_column_types(df)
        all_logs.extend(type_logs)

    object_cols = df.select_dtypes(include=["object"]).columns.tolist()
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()

    if options.get("trim", True) and object_cols:
        df, text_logs = standardize_text_columns(df, object_cols)
        all_logs.extend(text_logs)

    if options.get("duplicates", True):
        before = len(df)
        df = df.drop_duplicates()
        removed = before - len(df)
        if removed > 0:
            all_logs.append(f"👥 Removed {removed} duplicate rows")

    if options.get("missing", True):
        total_missing = int(df.isnull().sum().sum())

        for col in numeric_cols:
            if col in df.columns:
                strategy = options.get("fill_strategy", "zero")
                if strategy == "mean":
                    df[col] = df[col].fillna(df[col].mean())
                elif strategy == "median":
                    df[col] = df[col].fillna(df[col].median())
                else:
                    df[col] = df[col].fillna(0)

        object_cols_current = df.select_dtypes(include=["object"]).columns.tolist()
        for col in object_cols_current:
            df[col] = df[col].fillna("N/A")

        if total_missing > 0:
            all_logs.append(
                f"🩹 Filled {total_missing} missing values (numeric: {options.get('fill_strategy', 'zero')}, text: 'N/A')"
            )

    if options.get("outliers", False):
        numeric_cols_current = df.select_dtypes(include=[np.number]).columns.tolist()
        df, outlier_logs = remove_outliers_iqr(df, numeric_cols_current, multiplier=options.get("outlier_multiplier", 3.0))
        all_logs.extend(outlier_logs)

    if options.get("special", False):
        object_cols_current = df.select_dtypes(include=["object"]).columns.tolist()
        for col in object_cols_current:
            try:
                df[col] = df[col].astype(str).str.replace(r"[^\w\s\.\,\-\/\(\)]", "", regex=True)
            except Exception:
                pass
        all_logs.append("🧹 Special characters removed from text columns")

    numeric_cols_final = df.select_dtypes(include=[np.number]).columns.tolist()
    df, neg_logs = fix_negative_values(df, numeric_cols_final)
    all_logs.extend(neg_logs)

    clean_stats = {
        "original_rows": original_rows,
        "original_cols": original_cols,
        "final_rows": len(df),
        "final_cols": len(df.columns),
        "rows_removed": original_rows - len(df),
        "cols_removed": original_cols - len(df.columns),
    }
    return df, clean_stats, all_logs


# ==================== HELPERS ====================

def make_columns_unique(df):
    if df.columns.duplicated().any():
        cols = pd.Series(df.columns.astype(str))
        for dup in cols[cols.duplicated()].unique():
            count = 0
            for i, c in enumerate(cols):
                if c == dup:
                    if count > 0:
                        cols.iloc[i] = f"{c}_{count}"
                    count += 1
        df.columns = cols
    return df


def make_streamlit_safe(df):
    safe = make_columns_unique(df.copy())
    safe.columns = [str(c) for c in safe.columns]

    for col in safe.columns:
        s = safe[col]
        if pd.api.types.is_datetime64tz_dtype(s):
            safe[col] = s.dt.tz_localize(None)
        if s.dtype == "object":
            sample = s.dropna().head(50)
            if any(isinstance(v, (list, dict, set, tuple, bytes, bytearray)) for v in sample):
                safe[col] = s.astype(str)

    return safe


def load_csv(file_content, filename):
    encodings = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
    for enc in encodings:
        try:
            df = pd.read_csv(io.BytesIO(file_content), encoding=enc, on_bad_lines="skip")
            if not df.empty:
                df = df.dropna(how="all").dropna(axis=1, how="all")
                return make_columns_unique(df)
        except Exception:
            continue
    st.error("❌ Could not read CSV file")
    return None


def load_url_data(url):
    try:
        src = url.strip()
        if "docs.google.com/spreadsheets" in src and "/edit" in src:
            src = re.sub(r"/edit.*", "/export?format=csv", src)

        if src.lower().endswith((".xlsx", ".xls", ".xlsm")):
            df = pd.read_excel(src)
        else:
            df = pd.read_csv(src)

        df = df.dropna(how="all").dropna(axis=1, how="all")
        return make_columns_unique(df)
    except Exception as e:
        st.error(f"❌ Error loading URL data: {e}")
        return None


def load_excel(file_content, filename):
    try:
        excel_io = io.BytesIO(file_content)
        xls = pd.ExcelFile(excel_io)
        available = [s for s in xls.sheet_names if s.strip().upper() in TARGET_SHEETS]

        if not available:
            st.warning("⚠️ Required sheets (RM, PM, FG, SFG, SFG 1) were not found exactly. Showing all sheets.")
            available = xls.sheet_names

        selector_key = f"sheet_selector_{abs(hash(filename))}"
        chosen_sheet = st.selectbox("📑 Select Sheet", available, key=selector_key)

        parsed_df, header_idx, two_row = parse_excel_bytes(file_content, chosen_sheet)
        if parsed_df.empty:
            st.warning(f"⚠️ '{chosen_sheet}' sheet is empty after parsing.")
            return None

        parsed_df = make_columns_unique(parsed_df)
        st.session_state.source_sheet = chosen_sheet
        header_mode = "Two-row" if two_row else "Single-row"
        st.success(f"✅ Loaded Sheet: {chosen_sheet} | Header row detected at Excel row {header_idx + 1} ({header_mode})")
        return parsed_df
    except Exception as e:
        st.error(f"❌ Error loading Excel file: {e}")
        return None


def generate_sample_data():
    np.random.seed(42)
    n = 60
    data = {
        "ITEM": [f"Product_{i:03d}" for i in range(1, n + 1)],
        "CODE": [f"PRD-{i:03d}" for i in range(1, n + 1)],
        "CATEGORY": np.random.choice(["Raw Material", "Finished Goods", "Semi-Finished", "WIP"], n),
        "SUPPLIER": np.random.choice(["Supplier A", "Supplier B", "Supplier C", "Supplier D"], n),
        "SAFETY_STOCK": np.random.randint(100, 1000, n),
        "OPENING_STOCK": np.random.randint(500, 5000, n),
        "RECEIVED": np.random.randint(0, 2000, n),
        "ISSUED": np.random.randint(0, 3000, n),
        "CLOSING_STOCK": np.random.randint(0, 5000, n),
        "UNIT_PRICE": np.random.uniform(10, 500, n).round(2),
        "STATUS": np.random.choice(["Active", "Inactive", "Discontinued"], n),
        "LAST_UPDATED": pd.date_range("2024-01-01", periods=n, freq="5D").astype(str),
        "REORDER_POINT": np.random.randint(50, 500, n),
        "LEAD_TIME_DAYS": np.random.randint(1, 30, n),
    }
    df = pd.DataFrame(data)
    df["STOCK_VALUE"] = df["CLOSING_STOCK"] * df["UNIT_PRICE"]
    return df


def create_excel_export(df):
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Cleaned Data", index=False)
            wb = writer.book
            ws = writer.sheets["Cleaned Data"]

            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        output.seek(0)
        return output.getvalue()
    except Exception as e:
        st.error(f"Excel export error: {e}")
        return None


# ==================== CHART ENGINE ====================

def create_chart(chart_type, df, x_col, y_col, color_col, facet_col, facet_row, numeric_cols, text_cols, top_n=15):
    try:
        common_layout = dict(
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            height=500,
            font=dict(size=11),
            margin=dict(l=40, r=40, t=60, b=40),
        )

        if y_col and y_col in numeric_cols:
            plot_df = (
                df.nlargest(top_n, y_col)
                if chart_type not in [
                    "Line Chart",
                    "Area Chart",
                    "Scatter Plot",
                    "Histogram",
                    "Box Plot",
                    "Violin Plot",
                    "Heatmap",
                ]
                else df
            )
        else:
            plot_df = df.head(top_n)

        fc = facet_col if facet_col and facet_col != "None" else None
        fr = facet_row if facet_row and facet_row != "None" else None
        cc = color_col if color_col and color_col != "None" else None

        if chart_type == "Bar Chart":
            fig = px.bar(plot_df, x=x_col, y=y_col, color=cc, facet_col=fc, facet_row=fr, color_continuous_scale="Viridis", title=f"{y_col} by {x_col}")
        elif chart_type == "Horizontal Bar":
            fig = px.bar(plot_df, x=y_col, y=x_col, orientation="h", color=cc, facet_col=fc, color_continuous_scale="Viridis", title=f"Top {top_n} — {y_col}")
        elif chart_type == "Grouped Bar":
            fig = px.bar(plot_df, x=x_col, y=y_col, color=cc, barmode="group", facet_col=fc, title=f"Grouped: {y_col} by {x_col}")
        elif chart_type == "Stacked Bar":
            fig = px.bar(plot_df, x=x_col, y=y_col, color=cc, barmode="stack", facet_col=fc, title=f"Stacked: {y_col} by {x_col}")
        elif chart_type == "Pie Chart":
            fig = px.pie(plot_df, names=x_col, values=y_col, title=f"Distribution: {y_col}")
        elif chart_type == "Donut Chart":
            fig = px.pie(plot_df, names=x_col, values=y_col, hole=0.55, title=f"Distribution: {y_col}")
        elif chart_type == "Line Chart":
            fig = px.line(df, x=x_col, y=y_col, color=cc, facet_col=fc, facet_row=fr, markers=True, title=f"{y_col} Trend")
        elif chart_type == "Multi-Line":
            if len(numeric_cols) < 2:
                st.warning("Need ≥ 2 numeric columns for multi-line")
                return None
            fig = go.Figure()
            for nc in numeric_cols[:5]:
                fig.add_trace(go.Scatter(x=df[x_col], y=df[nc], name=nc, mode="lines+markers"))
            fig.update_layout(title=f"Multi-Line: {x_col}")
        elif chart_type == "Area Chart":
            fig = px.area(df, x=x_col, y=y_col, color=cc, facet_col=fc, title=f"{y_col} Area")
        elif chart_type == "Scatter Plot":
            fig = px.scatter(
                df,
                x=x_col,
                y=y_col,
                color=cc,
                facet_col=fc,
                facet_row=fr,
                trendline="ols" if options.get("trendline") else None,
                title=f"{x_col} vs {y_col}",
            )
        elif chart_type == "Bubble Chart":
            fig = px.scatter(df, x=x_col, y=y_col, size=y_col, color=cc, title=f"Bubble: {x_col} vs {y_col}")
        elif chart_type == "3D Scatter":
            if len(numeric_cols) < 3:
                st.warning("Need ≥ 3 numeric columns for 3D scatter")
                return None
            fig = px.scatter_3d(df, x=numeric_cols[0], y=numeric_cols[1], z=numeric_cols[2], color=cc, title="3D Scatter Plot")
        elif chart_type == "Histogram":
            fig = px.histogram(df, x=y_col, color=cc, facet_col=fc, nbins=30, marginal="box", title=f"Distribution: {y_col}")
        elif chart_type == "Box Plot":
            fig = px.box(df, x=x_col if x_col != y_col else None, y=y_col, color=cc, facet_col=fc, points="outliers", title=f"Box Plot: {y_col}")
        elif chart_type == "Violin Plot":
            fig = px.violin(df, x=x_col if x_col != y_col else None, y=y_col, color=cc, facet_col=fc, box=True, title=f"Violin: {y_col}")
        elif chart_type == "Heatmap":
            if len(numeric_cols) < 2:
                st.warning("Need ≥ 2 numeric columns for heatmap")
                return None
            corr = df[numeric_cols].corr()
            fig = px.imshow(corr, text_auto=".2f", aspect="auto", color_continuous_scale="RdBu_r", title="Correlation Heatmap")
        elif chart_type == "Pivot Heatmap":
            if not (x_col and y_col and cc):
                st.warning("Pivot Heatmap needs X, Y and Color By columns")
                return None
            pivot = df.pivot_table(index=x_col, columns=cc, values=y_col, aggfunc="sum", fill_value=0)
            fig = px.imshow(pivot, aspect="auto", color_continuous_scale="Blues", title=f"Pivot Heatmap: {y_col}")
        elif chart_type == "Treemap":
            path_cols = [c for c in [cc, x_col] if c]
            fig = px.treemap(plot_df, path=path_cols, values=y_col, title=f"Treemap: {y_col}")
        elif chart_type == "Sunburst":
            path_cols = [c for c in [x_col, cc] if c] or text_cols[:2]
            fig = px.sunburst(plot_df, path=path_cols, values=y_col, title="Sunburst Chart")
        elif chart_type == "Funnel Chart":
            fig = px.funnel(plot_df, x=y_col, y=x_col, title=f"Funnel: {y_col}")
        elif chart_type == "Waterfall":
            top_data = df.nlargest(10, y_col)
            fig = go.Figure(go.Waterfall(x=top_data[x_col].astype(str).tolist(), y=top_data[y_col].tolist(), textposition="outside"))
            fig.update_layout(title=f"Waterfall: {y_col}")
        elif chart_type == "Gauge":
            val = float(df[y_col].sum()) if y_col else 0
            max_val = val * 1.5 if y_col else 100
            fig = go.Figure(
                go.Indicator(
                    mode="gauge+number+delta",
                    value=val,
                    delta={"reference": val * 0.9},
                    gauge={
                        "axis": {"range": [0, max_val]},
                        "bar": {"color": "#667eea"},
                        "steps": [
                            {"range": [0, max_val * 0.5], "color": "#ff6b6b"},
                            {"range": [max_val * 0.5, max_val * 0.8], "color": "#feca57"},
                            {"range": [max_val * 0.8, max_val], "color": "#1abc9c"},
                        ],
                    },
                    title={"text": f"Total {y_col}"},
                )
            )
        elif chart_type == "Bullet Chart":
            fig = go.Figure()
            for _, row in plot_df.head(10).iterrows():
                fig.add_trace(
                    go.Indicator(
                        mode="number+gauge",
                        value=row[y_col],
                        domain={"x": [0, 1], "y": [0, 0.1]},
                        title={"text": str(row[x_col])[:20]},
                    )
                )
            fig.update_layout(title=f"Bullet Chart: {y_col}")
        elif chart_type == "Facet Bar Report":
            if not fc:
                st.warning("Select a 'Facet Column' for Facet Bar Report")
                return None
            fig = px.bar(plot_df, x=x_col, y=y_col, color=cc, facet_col=fc, facet_col_wrap=3, title=f"Facet Report: {y_col} by {x_col}", height=600)
            fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
        elif chart_type == "Facet Line Report":
            if not fc:
                st.warning("Select a 'Facet Column' for Facet Line Report")
                return None
            fig = px.line(df, x=x_col, y=y_col, color=cc, facet_col=fc, facet_col_wrap=3, markers=True, title=f"Facet Line Report: {y_col}", height=600)
            fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
        else:
            st.warning(f"Chart type '{chart_type}' not implemented")
            return None

        fig.update_layout(**common_layout)
        return fig
    except Exception as e:
        st.error(f"❌ Chart error ({chart_type}): {e}")
        return None


# ==================== SIDEBAR ====================

with st.sidebar:
    st.title("⚙️ Control Panel")
    st.markdown("---")

    st.subheader("🎨 Theme")
    theme = st.selectbox("Choose Theme", ["Professional Blue", "Dark Mode", "Ocean", "Sunset", "Forest", "Purple", "Mint", "Custom"])
    color_themes = {
        "Professional Blue": ("#667eea", "#764ba2", "#00cc00"),
        "Dark Mode": ("#1e1e1e", "#333333", "#4CAF50"),
        "Ocean": ("#0077be", "#00a8cc", "#7fdbff"),
        "Sunset": ("#ff6b6b", "#feca57", "#ff9ff3"),
        "Forest": ("#2d5016", "#73a942", "#aad576"),
        "Purple": ("#9b59b6", "#8e44ad", "#3498db"),
        "Mint": ("#1abc9c", "#16a085", "#27ae60"),
    }

    if theme == "Custom":
        primary_color = st.color_picker("Primary", "#667eea")
        secondary_color = st.color_picker("Secondary", "#764ba2")
        accent_color = st.color_picker("Accent", "#00cc00")
    else:
        primary_color, secondary_color, accent_color = color_themes[theme]

    st.session_state.theme_color = primary_color
    st.markdown("---")

    st.subheader("🔄 Auto Refresh")
    auto_refresh = st.checkbox("Enable Auto-Refresh", value=False)
    refresh_seconds = st.slider("Refresh interval (s)", 5, 300, 60)
    st.markdown("---")

    st.subheader("📁 Data Source")
    data_source = st.radio("Source", ["Upload File", "Live URL", "Sample Data"])
    st.markdown("---")

    st.subheader("🧹 Auto-Clean Options")
    remove_duplicates = st.checkbox("Remove Duplicates", value=True)
    fill_missing = st.checkbox("Fill Missing Values", value=True)
    fill_strategy = st.selectbox("Fill Strategy (numeric)", ["zero", "mean", "median"])
    trim_spaces = st.checkbox("Trim & Standardize Text", value=True)
    standardize_case = st.checkbox("Standardize Column Names", value=True)
    smart_types = st.checkbox("Smart Type Detection", value=True, help="Auto-detect currency, %, dates, booleans")
    remove_empty_rows = st.checkbox("Remove Empty Rows/Cols", value=True)
    remove_special = st.checkbox("Remove Special Characters", value=False)
    remove_outliers = st.checkbox("Remove Outliers (IQR×3)", value=False)
    outlier_mult = st.slider("IQR Multiplier", 1.5, 5.0, 3.0, 0.5) if remove_outliers else 3.0
    st.markdown("---")

    st.subheader("🖥️ Display")
    show_kpis = st.checkbox("Show KPIs", value=True)
    show_alerts = st.checkbox("Show Alerts", value=True)
    show_charts = st.checkbox("Show Charts", value=True)
    show_pivot = st.checkbox("Show Pivot", value=True)
    show_stats = st.checkbox("Show Statistics", value=True)
    show_raw = st.checkbox("Show Raw Data", value=True)
    st.markdown("---")

    st.info(f"🔄 Refreshes: {st.session_state.refresh_count}")


# ==================== CSS ====================

st.markdown(
    f"""
<style>
.stApp {{background: linear-gradient(135deg,{primary_color}15 0%,{secondary_color}15 100%)}}
.main-header {{
    background: linear-gradient(135deg,{primary_color} 0%,{secondary_color} 100%);
    padding:30px; border-radius:15px; color:white;
    text-align:center; margin-bottom:30px;
    box-shadow:0 10px 30px rgba(0,0,0,.15);
}}
.kpi-card {{
    background:white; padding:20px; border-radius:10px;
    border-left:5px solid {primary_color};
    box-shadow:0 5px 15px rgba(0,0,0,.08); margin-bottom:15px;
}}
.alert-critical {{
    background:linear-gradient(135deg,#ff6b6b,#ff4757);
    color:white; padding:15px; border-radius:10px; margin:8px 0; font-weight:bold;
}}
.alert-warning {{
    background:linear-gradient(135deg,#feca57,#ff9ff3);
    color:white; padding:15px; border-radius:10px; margin:8px 0; font-weight:bold;
}}
.alert-success {{
    background:linear-gradient(135deg,#1abc9c,#16a085);
    color:white; padding:15px; border-radius:10px; margin:8px 0; font-weight:bold;
}}
.clean-log {{
    background:#f8f9fa; border-left:4px solid {primary_color};
    padding:10px 15px; border-radius:5px; margin:4px 0; font-size:13px;
}}
</style>
""",
    unsafe_allow_html=True,
)


# ==================== HEADER ====================

st.markdown(
    """
<div class="main-header">
    <h1>📊 Ultimate Professional Dashboard</h1>
    <p>Deep Auto-Clean · Facet Reports · Live Data · Advanced Analytics</p>
</div>""",
    unsafe_allow_html=True,
)


# ==================== DATA LOADING ====================

df_raw = None
options = {
    "duplicates": remove_duplicates,
    "missing": fill_missing,
    "fill_strategy": fill_strategy,
    "trim": trim_spaces,
    "standardize": standardize_case,
    "smart_types": smart_types,
    "special": remove_special,
    "outliers": remove_outliers,
    "outlier_multiplier": outlier_mult,
    "empty_rows": remove_empty_rows,
    "empty_cols": remove_empty_rows,
}

if data_source == "Upload File":
    uploaded = st.file_uploader("📁 Upload CSV / Excel", type=["csv", "xlsx", "xlsm", "xls"])
    if uploaded:
        file_bytes = uploaded.read()
        if uploaded.name.lower().endswith(".csv"):
            df_raw = load_csv(file_bytes, uploaded.name)
        else:
            df_raw = load_excel(file_bytes, uploaded.name)

        if df_raw is not None:
            st.session_state.data = df_raw
            st.success(f"✅ Loaded {len(df_raw):,} rows × {len(df_raw.columns)} columns")

elif data_source == "Live URL":
    url = st.text_input("📎 Google Sheets / CSV URL")
    if url and st.button("🔄 Load"):
        with st.spinner("Loading..."):
            df_raw = load_url_data(url)
            if df_raw is not None:
                st.session_state.data = df_raw
                st.success(f"✅ Loaded {len(df_raw):,} rows")

elif data_source == "Sample Data":
    if st.button("📊 Generate Sample Data"):
        df_raw = generate_sample_data()
        st.session_state.data = df_raw
        st.success("✅ Sample data ready!")


# ==================== MAIN DASHBOARD ====================

if st.session_state.data is not None:
    df_raw = make_columns_unique(st.session_state.data.copy())

    with st.spinner("🧹 Running deep auto-clean..."):
        df, clean_stats, clean_log = deep_clean_data(df_raw.copy(), options)

    df = make_columns_unique(df)
    st.session_state.clean_data = df
    st.session_state.clean_log = clean_log

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    text_cols = df.select_dtypes(include=["object"]).columns.tolist()
    date_cols = df.select_dtypes(include=["datetime"]).columns.tolist()
    df_display = make_streamlit_safe(df)

    with st.expander("🧹 Deep-Clean Report", expanded=True):
        c1, c2, c3, c4, c5, c6 = st.columns(6)
        c1.metric("Original Rows", f"{clean_stats['original_rows']:,}")
        c2.metric("Cleaned Rows", f"{clean_stats['final_rows']:,}")
        c3.metric("Rows Removed", f"{clean_stats['rows_removed']:,}")
        c4.metric("Original Cols", clean_stats["original_cols"])
        c5.metric("Cleaned Cols", clean_stats["final_cols"])
        c6.metric("Actions Taken", len(clean_log))

        if st.session_state.source_sheet:
            st.caption(f"Source sheet: {st.session_state.source_sheet}")

        if clean_log:
            st.markdown("#### 📋 What Was Cleaned:")
            for entry in clean_log:
                st.markdown(f'<div class="clean-log">{entry}</div>', unsafe_allow_html=True)
        else:
            st.success("✅ Data was already clean — no changes needed!")

        type_df = pd.DataFrame(
            {
                "Column": df.columns,
                "Type": df.dtypes.astype(str).values,
                "Nulls": df.isnull().sum().values,
                "Unique": df.nunique().values,
                "Sample": [str(df[c].dropna().iloc[0])[:40] if df[c].dropna().shape[0] > 0 else "N/A" for c in df.columns],
            }
        )
        st.dataframe(make_streamlit_safe(type_df), use_container_width=True, hide_index=True)

    tabs = st.tabs(["📈 KPIs", "🚨 Alerts", "📊 Charts", "📐 Facet Report", "🔄 Pivot", "📉 Statistics", "🔍 Filter", "📋 Raw Data", "📥 Export"])

    with tabs[0]:
        if show_kpis:
            st.subheader("📈 Key Performance Indicators")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("📦 Records", f"{len(df):,}")
            m2.metric("📊 Columns", len(df.columns))
            m3.metric("🔢 Numeric Fields", len(numeric_cols))
            m4.metric("📝 Text Fields", len(text_cols))

            if numeric_cols:
                st.markdown("---")
                st.markdown("### 🔢 Numeric Column KPIs")
                ntabs = st.tabs([f"📊 {c[:18]}" for c in numeric_cols[:8]])
                for i, nt in enumerate(ntabs):
                    col_name = numeric_cols[i]
                    with nt:
                        a, b, c_, d, e, f_ = st.columns(6)
                        a.metric("Sum", f"{df[col_name].sum():,.0f}")
                        b.metric("Mean", f"{df[col_name].mean():,.2f}")
                        c_.metric("Median", f"{df[col_name].median():,.2f}")
                        d.metric("Max", f"{df[col_name].max():,.0f}")
                        e.metric("Min", f"{df[col_name].min():,.0f}")
                        f_.metric("Std", f"{df[col_name].std():,.2f}")

                        fig = px.histogram(df, x=col_name, nbins=25, marginal="box", title=f"Distribution — {col_name}")
                        st.plotly_chart(fig, use_container_width=True)

            if text_cols:
                st.markdown("---")
                st.markdown("### 📝 Categorical Breakdown")
                for col in text_cols[:4]:
                    with st.expander(f"📊 {col}"):
                        vc = df[col].value_counts()
                        c1, c2 = st.columns([1, 2])
                        with c1:
                            st.dataframe(make_streamlit_safe(vc.head(10).to_frame("COUNT")), use_container_width=True)
                        with c2:
                            fig = px.bar(x=vc.head(10).index, y=vc.head(10).values, labels={"x": col, "y": "Count"}, title=f"Top 10 — {col}")
                            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Enable KPIs in sidebar.")

    with tabs[1]:
        if show_alerts:
            st.subheader("🚨 Smart Alert System")
            found = False
            total_len = max(len(df), 1)

            for col in numeric_cols:
                zc = int((df[col] == 0).sum())
                pct = zc / total_len * 100
                if zc > 0 and pct > 20:
                    found = True
                    st.markdown(f'<div class="alert-warning">⚠️ <b>{col}</b>: {zc} zeros ({pct:.1f}%)</div>', unsafe_allow_html=True)

            for col in numeric_cols:
                neg = int((df[col] < 0).sum())
                if neg > 0:
                    found = True
                    st.markdown(f'<div class="alert-warning">⚠️ <b>{col}</b>: {neg} negative values</div>', unsafe_allow_html=True)

            if "CLOSING_STOCK" in df.columns and "SAFETY_STOCK" in df.columns:
                critical = df[df["CLOSING_STOCK"] < df["SAFETY_STOCK"]]
                if len(critical) > 0:
                    found = True
                    st.markdown(f'<div class="alert-critical">🔴 CRITICAL: {len(critical)} items below safety stock!</div>', unsafe_allow_html=True)
                    st.dataframe(make_streamlit_safe(critical), use_container_width=True)

            if "CLOSING_STOCK" in df.columns:
                oos = df[df["CLOSING_STOCK"] == 0]
                if len(oos) > 0:
                    found = True
                    st.markdown(f'<div class="alert-critical">🔴 OUT OF STOCK: {len(oos)} items!</div>', unsafe_allow_html=True)
                    st.dataframe(make_streamlit_safe(oos), use_container_width=True)

            if "ISSUED" in df.columns and "RECEIVED" in df.columns:
                dead = df[(df["ISSUED"] == 0) & (df["RECEIVED"] == 0)]
                if len(dead) > 0:
                    found = True
                    st.markdown(f'<div class="alert-warning">⚠️ DEAD STOCK: {len(dead)} items with no movement</div>', unsafe_allow_html=True)
                    st.dataframe(make_streamlit_safe(dead), use_container_width=True)

            null_pct = df.isnull().sum() / total_len * 100
            high_null = null_pct[null_pct > 30]
            if len(high_null) > 0:
                found = True
                for col, pct in high_null.items():
                    st.markdown(f'<div class="alert-warning">⚠️ <b>{col}</b>: {pct:.1f}% missing values</div>', unsafe_allow_html=True)

            if not found:
                st.markdown('<div class="alert-success">✅ No issues found — data looks healthy!</div>', unsafe_allow_html=True)
        else:
            st.info("Enable Alerts in sidebar.")

    with tabs[2]:
        if show_charts:
            st.subheader("📊 Advanced Visualizations")
            chart_types = [
                "Bar Chart",
                "Horizontal Bar",
                "Grouped Bar",
                "Stacked Bar",
                "Pie Chart",
                "Donut Chart",
                "Line Chart",
                "Multi-Line",
                "Area Chart",
                "Scatter Plot",
                "Bubble Chart",
                "3D Scatter",
                "Histogram",
                "Box Plot",
                "Violin Plot",
                "Heatmap",
                "Pivot Heatmap",
                "Treemap",
                "Sunburst",
                "Funnel Chart",
                "Waterfall",
                "Gauge",
                "Facet Bar Report",
                "Facet Line Report",
            ]

            ctrl, canvas = st.columns([1, 3])
            with ctrl:
                chart_type = st.selectbox("Chart Type", chart_types)
                x_col = st.selectbox("X-Axis", ["None"] + df.columns.tolist())
                y_col = st.selectbox("Y-Axis", ["None"] + numeric_cols)
                color_col = st.selectbox("Color By", ["None"] + df.columns.tolist())
                facet_col_s = st.selectbox("Facet Col", ["None"] + text_cols)
                facet_row_s = st.selectbox("Facet Row", ["None"] + text_cols)
                top_n = st.slider("Top N", 5, 100, 20)
                trendline = st.checkbox("Add Trendline", value=False)
                options["trendline"] = trendline

            with canvas:
                x_c = None if x_col == "None" else x_col
                y_c = None if y_col == "None" else y_col
                c_c = None if color_col == "None" else color_col
                fc = None if facet_col_s == "None" else facet_col_s
                fr = None if facet_row_s == "None" else facet_row_s

                if x_c and y_c:
                    fig = create_chart(chart_type, df, x_c, y_c, c_c, fc, fr, numeric_cols, text_cols, top_n)
                    if fig:
                        st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info("👆 Select X and Y axes")

            if numeric_cols and text_cols:
                st.markdown("---")
                st.markdown("### ⚡ Auto Quick-Charts")
                qc1, qc2, qc3 = st.columns(3)
                with qc1:
                    f = create_chart("Bar Chart", df, text_cols[0], numeric_cols[0], None, None, None, numeric_cols, text_cols, 10)
                    if f:
                        st.plotly_chart(f, use_container_width=True)
                with qc2:
                    f = create_chart("Donut Chart", df, text_cols[0], numeric_cols[0], None, None, None, numeric_cols, text_cols, 10)
                    if f:
                        st.plotly_chart(f, use_container_width=True)
                with qc3:
                    if len(numeric_cols) > 1:
                        f = create_chart("Scatter Plot", df, numeric_cols[0], numeric_cols[1], None, None, None, numeric_cols, text_cols)
                        if f:
                            st.plotly_chart(f, use_container_width=True)
        else:
            st.info("Enable Charts in sidebar.")

    with tabs[3]:
        st.subheader("📐 Facet Layout Report")
        st.info("💡 Facet reports split one chart into a grid of sub-charts — one panel per category.")

        if not numeric_cols or not text_cols:
            st.warning("Facet report needs at least one numeric and one text column.")
        else:
            fc1, fc2, fc3, fc4, fc5 = st.columns(5)
            with fc1:
                f_metric = st.selectbox("Metric (Y)", numeric_cols, key="fr_y")
            with fc2:
                f_xaxis = st.selectbox("X-Axis", df.columns.tolist(), key="fr_x")
            with fc3:
                f_split = st.selectbox("Split By (Facet)", text_cols, key="fr_split")
            with fc4:
                f_color = st.selectbox("Color By", ["None"] + text_cols, key="fr_color")
            with fc5:
                f_wrap = st.slider("Columns per Row", 1, 5, 3)
                f_chart_t = st.selectbox("Type", ["Bar", "Line", "Box", "Violin", "Histogram"])

            if st.button("🖼️ Generate Facet Report", type="primary"):
                color_arg = None if f_color == "None" else f_color
                try:
                    if f_chart_t == "Bar":
                        report_df = df.groupby([f_split, f_xaxis])[f_metric].sum().reset_index()
                        fig_r = px.bar(
                            report_df,
                            x=f_xaxis,
                            y=f_metric,
                            color=color_arg or f_metric,
                            facet_col=f_split,
                            facet_col_wrap=f_wrap,
                            title=f"Facet Bar Report: {f_metric} by {f_xaxis} | Split: {f_split}",
                            height=250 * ((df[f_split].nunique() // f_wrap) + 1),
                            color_continuous_scale="Viridis",
                        )
                    elif f_chart_t == "Line":
                        fig_r = px.line(
                            df,
                            x=f_xaxis,
                            y=f_metric,
                            color=color_arg,
                            facet_col=f_split,
                            facet_col_wrap=f_wrap,
                            markers=True,
                            title=f"Facet Line: {f_metric}",
                            height=250 * ((df[f_split].nunique() // f_wrap) + 1),
                        )
                    elif f_chart_t == "Box":
                        fig_r = px.box(
                            df,
                            x=f_xaxis,
                            y=f_metric,
                            color=color_arg,
                            facet_col=f_split,
                            facet_col_wrap=f_wrap,
                            title=f"Facet Box: {f_metric}",
                            height=250 * ((df[f_split].nunique() // f_wrap) + 1),
                        )
                    elif f_chart_t == "Violin":
                        fig_r = px.violin(
                            df,
                            x=f_xaxis,
                            y=f_metric,
                            color=color_arg,
                            facet_col=f_split,
                            facet_col_wrap=f_wrap,
                            box=True,
                            title=f"Facet Violin: {f_metric}",
                            height=250 * ((df[f_split].nunique() // f_wrap) + 1),
                        )
                    else:
                        fig_r = px.histogram(
                            df,
                            x=f_metric,
                            color=color_arg,
                            facet_col=f_split,
                            facet_col_wrap=f_wrap,
                            nbins=20,
                            title=f"Facet Histogram: {f_metric}",
                            height=250 * ((df[f_split].nunique() // f_wrap) + 1),
                        )

                    fig_r.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
                    fig_r.update_layout(plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                    st.plotly_chart(fig_r, use_container_width=True)
                except Exception as e:
                    st.error(f"❌ Facet report error: {e}")

            st.markdown("---")
            st.markdown("### 📋 Facet Summary Table")
            try:
                summary = df.groupby(f_split)[f_metric].agg(["sum", "mean", "min", "max", "count"]).round(2).reset_index()
                summary.columns = [f_split, "Sum", "Average", "Min", "Max", "Count"]
                st.dataframe(make_streamlit_safe(summary), use_container_width=True, hide_index=True)
            except Exception:
                pass

    with tabs[4]:
        if show_pivot:
            st.subheader("🔄 Pivot Table")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                pv_rows = st.multiselect("Rows", text_cols, default=text_cols[:1] if text_cols else [])
            with c2:
                pv_cols = st.multiselect("Columns", text_cols)
            with c3:
                pv_vals = st.multiselect("Values", numeric_cols, default=numeric_cols[:1] if numeric_cols else [])
            with c4:
                pv_agg = st.selectbox("Aggregation", ["sum", "mean", "count", "max", "min", "median"])

            if pv_rows and pv_vals:
                try:
                    pivot = pd.pivot_table(
                        df,
                        index=pv_rows,
                        columns=pv_cols if pv_cols else None,
                        values=pv_vals,
                        aggfunc=pv_agg,
                        fill_value=0,
                        margins=True,
                        margins_name="TOTAL",
                    )
                    pivot = make_columns_unique(pivot.reset_index())
                    st.dataframe(make_streamlit_safe(pivot), use_container_width=True, height=450)
                    st.download_button("📥 Download Pivot CSV", pivot.to_csv(index=False), f"pivot_{datetime.now():%Y%m%d_%H%M%S}.csv")
                except Exception as e:
                    st.error(f"Pivot error: {e}")
            else:
                st.info("Select Rows and Values to build pivot table.")
        else:
            st.info("Enable Pivot in sidebar.")

    with tabs[5]:
        if show_stats:
            st.subheader("📉 Statistical Analysis")

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("### Descriptive Statistics")
                st.dataframe(make_streamlit_safe(df.describe(include="all").T.round(3)), use_container_width=True)
            with c2:
                st.markdown("### Data Types")
                dt_df = pd.DataFrame(
                    {
                        "Type": df.dtypes.astype(str),
                        "Nulls": df.isnull().sum(),
                        "Null%": (df.isnull().sum() / max(len(df), 1) * 100).round(1),
                        "Unique": df.nunique(),
                    }
                )
                st.dataframe(make_streamlit_safe(dt_df), use_container_width=True)

            if numeric_cols:
                st.markdown("---")
                sel = st.selectbox("Deep-Dive Column", numeric_cols)
                c1, c2 = st.columns(2)
                with c1:
                    mode_v = df[sel].mode()
                    stats = {
                        "Count": df[sel].count(),
                        "Sum": f"{df[sel].sum():,.2f}",
                        "Mean": f"{df[sel].mean():,.2f}",
                        "Median": f"{df[sel].median():,.2f}",
                        "Mode": f"{mode_v.iloc[0]:,.2f}" if len(mode_v) else "N/A",
                        "Std Dev": f"{df[sel].std():,.2f}",
                        "Variance": f"{df[sel].var():,.2f}",
                        "Min": f"{df[sel].min():,.2f}",
                        "Q1 (25%)": f"{df[sel].quantile(.25):,.2f}",
                        "Q3 (75%)": f"{df[sel].quantile(.75):,.2f}",
                        "Max": f"{df[sel].max():,.2f}",
                        "Range": f"{df[sel].max() - df[sel].min():,.2f}",
                        "IQR": f"{df[sel].quantile(.75) - df[sel].quantile(.25):,.2f}",
                        "Skewness": f"{df[sel].skew():,.4f}",
                        "Kurtosis": f"{df[sel].kurtosis():,.4f}",
                    }
                    st.dataframe(make_streamlit_safe(pd.DataFrame(stats.items(), columns=["Metric", "Value"])), use_container_width=True, hide_index=True)
                with c2:
                    fig = px.histogram(df, x=sel, nbins=30, marginal="box", title=f"Distribution — {sel}")
                    st.plotly_chart(fig, use_container_width=True)

            if len(numeric_cols) > 1:
                st.markdown("---")
                st.markdown("### 🔗 Correlation Matrix")
                corr = df[numeric_cols].corr()
                fig = px.imshow(corr, text_auto=".2f", aspect="auto", color_continuous_scale="RdBu_r", title="Correlation Matrix")
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Enable Statistics in sidebar.")

    with tabs[6]:
        st.subheader("🔍 Filter & Query")

        search = st.text_input("🔍 Global Search")
        n_filt = st.slider("Number of filters", 1, 6, 2)

        filtered = df.copy()
        for i in range(n_filt):
            a, b = st.columns(2)
            with a:
                fc_sel = st.selectbox(f"Column {i + 1}", df.columns.tolist(), key=f"fc_{i}")
            with b:
                if fc_sel in numeric_cols:
                    lo = float(df[fc_sel].min())
                    hi = float(df[fc_sel].max())
                    rng = st.slider(f"{fc_sel} range", lo, hi, (lo, hi), key=f"fr_{i}")
                    filtered = filtered[(filtered[fc_sel] >= rng[0]) & (filtered[fc_sel] <= rng[1])]
                else:
                    opts = sorted(df[fc_sel].astype(str).unique())
                    sel_ = st.multiselect(f"{fc_sel} values", opts, default=opts[: min(5, len(opts))], key=f"fv_{i}")
                    if sel_:
                        filtered = filtered[filtered[fc_sel].astype(str).isin(sel_)]

        if search:
            mask = filtered.astype(str).apply(lambda x: x.str.contains(search, case=False, na=False)).any(axis=1)
            filtered = filtered[mask]

        base_len = max(len(df), 1)
        st.markdown(f"### Results: **{len(filtered):,}** rows ({len(filtered) / base_len * 100:.1f}% of {len(df):,})")
        st.dataframe(make_streamlit_safe(filtered), use_container_width=True, height=450)

        d1, d2, d3 = st.columns(3)
        with d1:
            st.download_button("📥 CSV", filtered.to_csv(index=False), f"filtered_{datetime.now():%Y%m%d_%H%M%S}.csv")
        with d2:
            st.download_button("📋 JSON", filtered.to_json(orient="records", indent=2), f"filtered_{datetime.now():%Y%m%d_%H%M%S}.json")
        with d3:
            xl = create_excel_export(filtered)
            if xl:
                st.download_button("📊 Excel", xl, f"filtered_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

    with tabs[7]:
        if show_raw:
            st.subheader("📋 Raw Cleaned Data")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Rows", f"{len(df):,}")
            m2.metric("Columns", len(df.columns))
            m3.metric("Nulls", int(df.isnull().sum().sum()))
            m4.metric("Dupes", int(df.duplicated().sum()))

            q = st.text_input("🔍 Search data")
            disp = df
            if q:
                mask = df.astype(str).apply(lambda x: x.str.contains(q, case=False, na=False)).any(axis=1)
                disp = df[mask]

            st.dataframe(make_streamlit_safe(disp), use_container_width=True, height=550)
        else:
            st.info("Enable Raw Data in sidebar.")

    with tabs[8]:
        st.subheader("📥 Export Options")

        e1, e2, e3, e4 = st.columns(4)
        with e1:
            st.download_button("📄 CSV", df.to_csv(index=False), f"data_{datetime.now():%Y%m%d_%H%M%S}.csv", use_container_width=True)
        with e2:
            st.download_button("📋 JSON", df.to_json(orient="records", indent=2), f"data_{datetime.now():%Y%m%d_%H%M%S}.json", use_container_width=True)
        with e3:
            xl = create_excel_export(df)
            if xl:
                st.download_button("📊 Excel", xl, f"data_{datetime.now():%Y%m%d_%H%M%S}.xlsx", use_container_width=True)
        with e4:
            try:
                st.download_button("⚡ Parquet", df.to_parquet(index=False), f"data_{datetime.now():%Y%m%d_%H%M%S}.parquet", use_container_width=True)
            except Exception as e:
                st.warning(f"Parquet unavailable: {e}")

    if auto_refresh:
        st.session_state.refresh_count += 1
        st.info(f"🔄 Auto-refreshing in {refresh_seconds}s...")
        time.sleep(refresh_seconds)
        st.rerun()

else:
    st.markdown(
        """
    <div class="main-header">
        <h2>👋 Welcome! Upload data to begin.</h2>
        <p>CSV · Excel · Google Sheets · Sample Data</p>
    </div>""",
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    cards = [
        (
            "🧹 Deep Auto-Clean",
            [
                "Currency & % detection",
                "Date parsing (13 formats)",
                "Boolean conversion",
                "Outlier removal (IQR)",
                "Duplicate & null handling",
            ],
        ),
        (
            "📐 Facet Reports",
            [
                "Bar / Line / Box / Violin",
                "Histogram facets",
                "Configurable grid wrap",
                "Color by category",
                "Summary table",
            ],
        ),
        (
            "📊 24 Chart Types",
            [
                "Grouped & Stacked Bar",
                "3D Scatter",
                "Gauge & Bullet",
                "Treemap & Sunburst",
                "Pivot Heatmap",
            ],
        ),
    ]

    for col, (title, items) in zip([c1, c2, c3], cards):
        with col:
            li = "".join(f"<li>✅ {i}</li>" for i in items)
            st.markdown(f'<div class="kpi-card"><h3>{title}</h3><ul>{li}</ul></div>', unsafe_allow_html=True)
